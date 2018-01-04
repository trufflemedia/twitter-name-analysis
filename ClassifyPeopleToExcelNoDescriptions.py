# 2017 12 29 John Blue fixed output file to have correct underscore reading (minor edit)
# 2017 09 23 John Blue took out printing descriptions, noted below
# 2017 09 07 John Blue this can be a memory hog on larger input sheets 
#  like new_users__FarmJournal_1504207365027.xlsx
# 2017 09 06 John Blue classify people in a spreadsheet
# see info at http://zetcode.com/articles/openpyxl/
# see also https://openpyxl.readthedocs.io/en/default/tutorial.html#playing-with-data
# updated to send data to an Excel file instead of a text file

# The input file must have screen_name in column 2, description in column 5, 
# and source in column 26.

# There is text output that can be piped to another file or ignored. It is used here as a 
# status and debugging tool.

# use
# python ClassifyPeopleToExcelNoDescriptions.py new_excel-file.xlsx terms.txt> out_no_decscript_new_excel-file.txt

# This will create a file called out_no_decscript_new_excel-file.xlsx in which all 
# the terms are column names and each row is the Twitter name.

# Example output on Google docs at
# https://drive.google.com/open?id=1kJPnDu2RjFXBPydmCnCQxRwxH9-Re8gERO98KFANhVU

from openpyxl import Workbook
from openpyxl import load_workbook
#from __future__ import print_function

import sys

# output to an already open worksheet a value at row, column.
# Then advance the column and return it back to caller


def OutToSheet(ASheet, Avalue, Arow, Acol):
    # output row number should mirror input row number
    ASheet.cell(row=Arow, column=Acol).value = Avalue
    Acol = Acol + 1  # move to next column
    return Acol


excelfilename = sys.argv[1]  # get source Excel filename
termfile = sys.argv[2]  # get terms filename

terms = [""]
# read terms into list, make all lower case without carriage return
# from
# https://stackoverflow.com/questions/19062574/python-read-file-into-list-strip-newlines
with open(termfile) as temp_file:
    terms = [line.rstrip('\n').lower() for line in temp_file]

# print terms

# setup the output Excel file for classification info
out_wb = Workbook()
out_ws = out_wb.active  # active worksheet
out_ws.title = "Classification Data"  # name the worksheet

out_filename = "out_no_decscript_" + excelfilename

# setup initial row in sheet with column header
# Important: first row is a header
out_acol = 1
print '{}|{}|'.format("screen_name", "source"),
out_acol = OutToSheet(out_ws, "screen_name", 1, out_acol)
out_acol = OutToSheet(out_ws, "source", 1, out_acol)

for t in terms:  # loop through terms to setup initial row in sheet
    print '{}|'.format(t),
    out_acol = OutToSheet(out_ws, t, 1, out_acol)
# end loop through terms

# description 
#print "targetstring"  # final print needed to end line correctly
print " "  # final print needed to end line correctly

# description 
#out_acol = OutToSheet(out_ws, "targetstring", 1, out_acol)

# setup input Excel file and worksheet
in_wb = load_workbook(excelfilename)  # workbook from file
in_ws = in_wb.worksheets[0]  # first worksheet from workbook

# identifiers for specific columns in input worksheet to use for classification
screen_name = 2
description = 5
source = 26

rmax = in_ws.max_row + 1  # need to get that last row; this seemed to work
# The first row of the input file is just a basic header and we are not
# using it
rmin = in_ws.min_row + 1

targetcolumn = in_ws.max_column + 1

# http://nullege.com/codes/search/openpyxl.worksheet.Worksheet.cell
for arow in range(rmin, rmax):  # begin loop through input file rows
    out_acol = 1  # Start on the left side of sheet output sheet
    row_screen_name = in_ws.cell(row=arow, column=screen_name).value
    if not row_screen_name:
        row_screen_name = "screen-null"

    row_description = in_ws.cell(row=arow, column=description).value
    if not row_description:
        row_description = "decription-null"

    row_source = in_ws.cell(row=arow, column=source).value
    if not row_source:
        row_source = "source-null"

    targetstring = (
        row_screen_name +
        " " +
        row_description).encode('utf-8').lower()

# output initial info on the account
    print '{}|{}|'.format(row_screen_name, row_source),
    out_acol = OutToSheet(out_ws, row_screen_name, arow, out_acol)
    out_acol = OutToSheet(out_ws, row_source, arow, out_acol)

    for t in terms:  # loop through terms

        if t in targetstring:
            print '{}|'.format('1'), #use 1 instead of t to get a numeric value in worksheet
            out_acol = OutToSheet(out_ws, 1, arow, out_acol) #use 1 instead of t to get a numeric value in worksheet
        else:
            print '{}|'.format('0'),
            out_acol = OutToSheet(out_ws, 0, arow, out_acol)
    # end loop through terms

#   print #targetstring # this is the final close of the row's info
# print targetstring #.replace('\n', ' ').replace('\r', '') # this is the
# final close of the row's info
    # this is the final close of the row's info, removing CR,LF,FF

    print " " # final print
# description 
#   print targetstring.replace('\n', ' ').replace('\r', '').replace('\f', '')

# description 
#    out_acol = OutToSheet(
#       out_ws,
#       targetstring.replace(
#            '\n',
#           ' ').replace(
#           '\r',
#           '').replace(
#               '\f',
#               ''),
#       arow,
#       out_acol)

# end loop through input file rows

#       ws.cell(row=arow, column=targetcolumn).value= excelfilename
out_wb.save(out_filename)  # save file
