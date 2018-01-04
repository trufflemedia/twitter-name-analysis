# 2017 09 06 John Blue Get Excel file name, insert column, 
# add the Excel's filename into all the cells of that new column. 
# This is used to help know where data originated.
# see info at http://zetcode.com/articles/openpyxl/


# use
# python AddColumnFilename.py excel-file.xlsx
# 
# This will create a file called new_ex/cel-file.xlsx in which the last
# column of each row contains "excel-file.xlsx"

from openpyxl import Workbook
from openpyxl import load_workbook

import sys

excelfilename=sys.argv[1]
wb=load_workbook(excelfilename) #workbook from file
ws=wb.worksheets[0] #first worksheet from workbook

rmax = ws.max_row
rmin = ws.min_row

targetcolumn = ws.max_column+1

# http://nullege.com/codes/search/openpyxl.worksheet.Worksheet.cell
for arow in range(rmin, rmax+1):
   ws.cell(row=arow, column=targetcolumn).value= excelfilename
   
newsavefile= "new_" + excelfilename
wb.save(newsavefile)
